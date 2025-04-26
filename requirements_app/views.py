from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth import logout
from django.http import HttpResponse, JsonResponse
from django.contrib import messages
from django.utils import timezone
from .models import User, RequirementForm, FormSection, FormQuestion, QuestionResponse
from .forms import RequirementFormForm, UserRegistrationForm, UserUpdateForm, DynamicForm, FormQuestionForm, FormSectionForm
from .utils import generate_pdf, generate_user_list_pdf, get_static_file_as_data_uri
import json
from django.db.models import Count, Avg, Q
from collections import defaultdict
from django.views.decorators.http import require_POST
from django.contrib.auth import authenticate, login
from .forms import LoginForm  
import os
import base64
from django.conf import settings
from django.core.files.base import ContentFile
from django.core.files.uploadedfile import UploadedFile
import tempfile
import pickle
from django.db.models.fields.files import FieldFile
from django.template.loader import render_to_string
from datetime import datetime
from weasyprint import HTML, CSS
from django.core.paginator import Paginator
from django.db.models import Count, IntegerField, Q
from django.db.models.functions import Coalesce
from django.db.models import Value as V
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from io import BytesIO
import datetime


def is_admin(user):
    return user.is_authenticated and user.role == 'admin'

@login_required
def dashboard(request):
    if request.user.role == 'admin':
        return redirect('admin_dashboard')
    
    # Clean up duplicate drafts if they exist
    cleanup_duplicate_drafts(request.user)
    
    # Get count of forms
    user_forms_count = RequirementForm.objects.filter(user=request.user).count()
    submitted_forms_count = RequirementForm.objects.filter(user=request.user, status='submitted').count()
    
    # Get draft forms for the user
    draft_forms = RequirementForm.objects.filter(user=request.user, status='draft').order_by('-updated_at')
    
    context = {
        'user_forms_count': user_forms_count,
        'submitted_forms_count': submitted_forms_count,
        'draft_forms': draft_forms,
    }
    return render(request, 'requirements_app/dashboard.html', context)

def cleanup_duplicate_drafts(user):
    """Find and merge duplicate drafts for a user"""
    # Find drafts with duplicate process names
    drafts = RequirementForm.objects.filter(user=user, status='draft')
    
    # Get process names with duplicates
    duplicate_process_names = drafts.values('process_name').annotate(
        count=Count('id')
    ).filter(count__gt=1).values_list('process_name', flat=True)
    
    for process_name in duplicate_process_names:
        # Get all drafts for this process name
        process_drafts = drafts.filter(process_name=process_name).order_by('-updated_at')
        
        if process_drafts.count() > 1:
            # Keep the most recently updated one
            most_recent = process_drafts.first()
            
            # Delete the rest
            for old_draft in process_drafts[1:]:
                old_draft.delete()

@login_required
def form_view(request, form_id=None):
    # Check if coming from review page with edit request
    edit_after_review = request.session.pop('edit_after_review', False)
    form_data_from_session = None
    
    if edit_after_review and 'form_data' in request.session:
        form_data_from_session = request.session['form_data']
        question_responses = request.session.get('question_responses', {})

    # If form_id is provided, get the specific form to edit (draft)
    if form_id:
        instance = get_object_or_404(RequirementForm, id=form_id, user=request.user, status='draft')
        # Convert the JSONField to a JSON string for the template
        instance_process_steps_detail = json.dumps(instance.process_steps_detail)
    else:
        instance = None
        instance_process_steps_detail = '[]'
    
    # Override instance_process_steps_detail if we have data from session
    if form_data_from_session and 'process_steps_detail' in form_data_from_session:
        # No need to convert to JSON string here since we'll do that in the template
        instance_process_steps_detail = json.dumps(form_data_from_session['process_steps_detail'])
        
    # Get all active form sections with their questions
    active_sections = FormSection.objects.filter(is_active=True).prefetch_related('questions').order_by('order')

    if request.method == 'POST':
        # Determine if this is a draft save
        is_draft_save = 'save_draft' in request.POST
        is_auto_save = 'auto_save' in request.POST
        # For draft saves, bypass form validation completely
        if is_draft_save:
            # For auto-save, try to find an existing draft with the same process name
            process_name = request.POST.get('process_name', '').strip()
            if is_auto_save and process_name:
                existing_drafts = RequirementForm.objects.filter(
                    user=request.user,
                    status='draft',
                    process_name=process_name
                )

                # If not the current instance and we found other drafts with same name,
                # use the most recently updated one
                if not instance and existing_drafts.exists():
                    instance = existing_drafts.order_by('-updated_at').first()
                    form_id = instance.id

            # Determine if we're updating an existing draft or creating a new one
            if instance:
                requirement_form = instance
            else:
                requirement_form = RequirementForm(user=request.user)
            
            # Set the status to draft
            requirement_form.status = 'draft'
            
            # Directly update fields from POST data
            # For text fields
            for field_name in ['process_name', 'process_description', 'expected_features', 
                            'internal_connectivity_details', 'external_connectivity_details',
                            'expected_reports', 'expected_analysis']:
                if field_name in request.POST:
                    setattr(requirement_form, field_name, request.POST.get(field_name, ''))
            
            # For integer fields
            for field_name in ['time_taken', 'people_involved', 'process_steps']:
                try:
                    value = request.POST.get(field_name, '')
                    if value and value.strip():
                        setattr(requirement_form, field_name, int(value))
                    else:
                        # Set default value of 0 to avoid NULL errors
                        setattr(requirement_form, field_name, 0)
                except (ValueError, TypeError):
                    # Set default value
                    setattr(requirement_form, field_name, 0)
                    pass
            
            # For choice fields
            for field_name in ['internal_connectivity', 'external_connectivity']:
                if field_name in request.POST:
                    setattr(requirement_form, field_name, request.POST.get(field_name, 'no'))
            
            # Process steps detail
            process_steps_detail = request.POST.get('process_steps_detail', '[]')
            try:
                requirement_form.process_steps_detail = json.loads(process_steps_detail)
            except json.JSONDecodeError:
                requirement_form.process_steps_detail = []
            
            # File uploads - only handle for non-auto-save to avoid partial uploads
            if not is_auto_save:
                if 'flowchart' in request.FILES:
                    requirement_form.flowchart = request.FILES['flowchart']
                
                if 'attachment' in request.FILES:
                    requirement_form.attachment = request.FILES['attachment']
            
            # Save the form (no validation)
            requirement_form.save()
            
            # Save question responses
            for field_name, value in request.POST.items():
                if field_name.startswith('question_'):
                    try:
                        question_id = int(field_name.split('_')[1])
                        question = FormQuestion.objects.get(id=question_id)
                        QuestionResponse.objects.update_or_create(
                            form=requirement_form,
                            question=question,
                            defaults={'response_text': value if value is not None else ''}
                        )
                    except (ValueError, FormQuestion.DoesNotExist):
                        pass
            
            # Check if it's an AJAX request
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'form_id':requirement_form.id})
                
            messages.success(request, "Form saved as draft successfully.")
            return redirect('dashboard')
        
        else:
            # Regular form processing with validation for non-draft
            basic_form = RequirementFormForm(request.POST, request.FILES, instance=instance, is_draft=False)
            dynamic_form = DynamicForm(request.POST, sections=active_sections, instance=instance, is_draft=False)
            
            # Process name error check (only for final submissions)
            process_name_error = None
            process_name = request.POST.get('process_name')
            query = RequirementForm.objects.filter(user=request.user, process_name=process_name, status='submitted')
            if instance:
                query = query.exclude(pk=instance.pk)
            if query.exists():
                process_name_error = f"You have already submitted a form for '{process_name}'. Please use a different process name."
                messages.error(request, process_name_error)
            
            elif basic_form.is_valid() and dynamic_form.is_valid():
                # Regular form processing for valid forms
                requirement_form = basic_form.save(commit=False)
                requirement_form.user = request.user
                
                # Process steps handling
                process_steps_detail = request.POST.get('process_steps_detail', '[]')
                try:
                    steps_data = json.loads(process_steps_detail)
                    valid_steps = []
                    for i, step in enumerate(steps_data):
                        if isinstance(step, dict) and 'description' in step and step['description'].strip():
                            step['number'] = step.get('number', i + 1)
                            valid_steps.append(step)
                        elif isinstance(step, str) and step.strip():
                            valid_steps.append({
                                'number': i + 1,
                                'description': step.strip()
                            })
                    
                    requirement_form.process_steps_detail = valid_steps
                    requirement_form.process_steps = len(valid_steps)
                except json.JSONDecodeError:
                    requirement_form.process_steps_detail = []
                    requirement_form.process_steps = 0
                
                # Handle review option
                if 'review' in request.POST:
                    # Store form data in session
                    form_data = {}

                    # Copy basic form data from POST (avoiding file fields completely)
                    for key in request.POST:
                        if key not in ['csrfmiddlewaretoken', 'save_draft', 'review', 'submit']:
                            form_data[key] = request.POST[key]
                    
                    # Store the processed steps data correctly
                    try:
                        steps_data = json.loads(process_steps_detail)
                        form_data['process_steps_detail'] = steps_data
                    except json.JSONDecodeError:
                        form_data['process_steps_detail'] = []
                    
                    # Handle file flags
                    form_data['has_flowchart'] = bool(request.FILES.get('flowchart')) or (instance and bool(instance.flowchart))
                    form_data['has_attachment'] = bool(request.FILES.get('attachment')) or (instance and bool(instance.attachment))
                
                    # Store form_id if we're editing an existing draft
                    if instance:
                        form_data['form_id'] = instance.id
                        
                        # Set file flags based on existing files
                        if not form_data['has_flowchart'] and instance.flowchart:
                            form_data['has_flowchart'] = True
                        if not form_data['has_attachment'] and instance.attachment:
                            form_data['has_attachment'] = True
                    
                    # Store dynamic form responses
                    question_responses = {}
                    for key, value in request.POST.items():
                        if key.startswith('question_'):
                            question_id = key.split('_')[1]
                            
                            # Convert list values to string for checkboxes
                            if isinstance(value, list):
                                value = ', '.join(value)
                                
                            question_responses[question_id] = value if value is not None else ''
                    
                    # Store files temporarily
                    temp_files = {}
                    if 'flowchart' in request.FILES:
                        temp_file = tempfile.NamedTemporaryFile(delete=False)
                        for chunk in request.FILES['flowchart'].chunks():
                            temp_file.write(chunk)
                        temp_file.close()
                        temp_files['flowchart'] = temp_file.name
                    
                    if 'attachment' in request.FILES:
                        temp_file = tempfile.NamedTemporaryFile(delete=False)
                        for chunk in request.FILES['attachment'].chunks():
                            temp_file.write(chunk)
                        temp_file.close()
                        temp_files['attachment'] = temp_file.name
                    
                    # Store everything in session
                    request.session['form_data'] = form_data
                    request.session['question_responses'] = question_responses
                    request.session['temp_files'] = temp_files
                    request.session['review_timestamp'] = timezone.now().isoformat()
                    
                    # Redirect to review page
                    return redirect('review_form')
            else:
                # Display form errors
                for field, errors in basic_form.errors.items():
                    for error in errors:
                        messages.error(request, f"{field}: {error}")
                
                for field, errors in dynamic_form.errors.items():
                    for error in errors:
                        messages.error(request, f"{field}: {error}")
    
    else:
        # Initial form load
        if form_data_from_session:
            # Use data from session for initial form values
            initial_data = {k: v for k, v in form_data_from_session.items() 
                            if k not in ['csrfmiddlewaretoken', 'process_steps_detail', 'has_flowchart', 'has_attachment']}
            
            # Special handling for some fields if needed
            if 'process_steps_detail' in form_data_from_session:
                instance_process_steps_detail = json.dumps(form_data_from_session['process_steps_detail'])
            
            basic_form = RequirementFormForm(initial=initial_data, instance=instance)
            
            # Create dynamic form with stored responses
            dynamic_form = DynamicForm(
                initial=question_responses,
                sections=active_sections, 
                instance=instance
            )
        else:
            basic_form = RequirementFormForm(instance=instance)
            dynamic_form = DynamicForm(sections=active_sections, instance=instance)
    
    # Group questions by section
    sections_with_questions = []
    for section in active_sections:
        questions = []
        for question in section.questions.filter(is_active=True).order_by('order'):
            field_name = f"question_{question.id}"
            if field_name in dynamic_form.fields:
                questions.append({
                    'question': question,
                    'field': dynamic_form[field_name]
                })
        
        if questions:
            sections_with_questions.append({
                'section': section,
                'questions': questions
            })
    
    context = {
        'basic_form': basic_form,
        'dynamic_form': dynamic_form,
        'sections_with_questions': sections_with_questions,
        'is_draft': instance is not None,
        'form_id': form_id,
        'process_name_error': process_name_error if 'process_name_error' in locals() else None,
        'instance_process_steps_detail': instance_process_steps_detail,
        'form_data': form_data_from_session,  # Add this to pass session data to template
    }
    
    return render(request, 'requirements_app/form.html', context)

@login_required
def view_submissions(request):
    forms = RequirementForm.objects.filter(user=request.user, status='submitted').order_by('-submitted_at')
    
    # Add attachment flags for each form
    for form in forms:
        form.has_attachment = bool(form.attachment)
        form.has_flowchart = bool(form.flowchart)
    
    context = {
        'forms': forms,
    }
    return render(request, 'requirements_app/view_submissions.html', context)

# Update view_responses for admin to see process name and user ID
@user_passes_test(is_admin)
def view_responses(request):
    forms = RequirementForm.objects.filter(status='submitted').order_by('-submitted_at')
    
    # Filter by process name if provided
    process_filter = request.GET.get('process')
    if process_filter:
        forms = forms.filter(process_name__icontains=process_filter)
    
    # Filter by user ID if provided
    user_filter = request.GET.get('user_id')
    if user_filter:
        forms = forms.filter(user__username__icontains=user_filter)
    
    # Check attachment existence for each form
    for form in forms:
        form.has_attachment = bool(form.attachment)
        form.has_flowchart = bool(form.flowchart)
    
    # Get unique process names for filtering
    all_process_names = RequirementForm.objects.filter(status='submitted').values_list('process_name', flat=True).distinct()
    
    context = {
        'forms': forms,
        'all_process_names': all_process_names,
        'process_filter': process_filter,
        'user_filter': user_filter,
    }
    return render(request, 'requirements_app/view_responses.html', context)

@login_required
def view_form(request, form_id=None):
    if form_id:
        form = get_object_or_404(RequirementForm, id=form_id, user=request.user)
    else:
        # If no ID provided, redirect to submissions page
        return redirect('view_submissions')
    
    # Get all sections with questions
    sections = FormSection.objects.filter(is_active=True).order_by('order')
    
    # Create a data structure to hold sections with their answers
    sections_with_answers = []
    
    for section in sections:
        section_answers = []
        for question in section.questions.filter(is_active=True):
            try:
                answer = QuestionResponse.objects.get(form=form, question=question)
                section_answers.append({
                    'question': question,
                    'answer_text': answer.response_text
                })
            except QuestionResponse.DoesNotExist:
                # Skip questions with no answers
                continue
                
        # Only add sections that have at least one answer
        if section_answers:
            sections_with_answers.append({
                'section': section,
                'answers': section_answers
            })
    
    context = {
        'form': form,
        'sections_with_answers': sections_with_answers,
        # data is available for template
        'user': request.user
    }
    return render(request, 'requirements_app/view_form.html', context)

@user_passes_test(is_admin)
def view_user_form(request, form_id):
    form = get_object_or_404(RequirementForm, pk=form_id)
    
    # Get all sections with questions
    sections = FormSection.objects.filter(is_active=True).order_by('order')
    
    # Create a data structure to hold sections with their answers
    sections_with_answers = []
    
    for section in sections:
        section_answers = []
        for question in section.questions.filter(is_active=True):
            try:
                answer = QuestionResponse.objects.get(form=form, question=question)
                section_answers.append({
                    'question': question,
                    'answer_text': answer.response_text
                })
            except QuestionResponse.DoesNotExist:
                # Skip questions with no answers
                continue
                
        # Only add sections that have at least one answer
        if section_answers:
            sections_with_answers.append({
                'section': section,
                'answers': section_answers
            })
    
    context = {
        'form': form,
        'sections_with_answers': sections_with_answers
    }
    return render(request, 'requirements_app/view_user_form.html', context)

@login_required
def download_pdf(request, form_id):
    form = get_object_or_404(RequirementForm, id=form_id, user=request.user)
    
    # Get all sections with questions and answers (same as view_form)
    sections = FormSection.objects.filter(is_active=True).order_by('order')
    sections_with_answers = []
    for section in sections:
        section_answers = []
        for question in section.questions.filter(is_active=True):
            try:
                answer = QuestionResponse.objects.get(form=form, question=question)
                section_answers.append({
                    'question': question,
                    'answer_text': answer.response_text
                })
            except QuestionResponse.DoesNotExist:
                continue
        if section_answers:
            sections_with_answers.append({
                'section': section,
                'answers': section_answers
            })

    # Generate the signature data URI
    signature_data_uri = None
    if request.user.signature:  # Check if the user has a signature
        signature_path = os.path.join(settings.MEDIA_ROOT, request.user.signature.name)
        if os.path.exists(signature_path):
            with open(signature_path, 'rb') as image_file:
                signature_data = base64.b64encode(image_file.read()).decode('utf-8')
                signature_data_uri = f"data:image/png;base64,{signature_data}"  # Adjust MIME type if needed
        else:
            print(f"Signature file not found at: {signature_path}")

    # Prepare context for the template
    context = {
        'form': form,
        'sections_with_answers': sections_with_answers,
        'user': request.user,  # Match view_form.html
        'current_date': timezone.now(),
        'signature_data_uri': signature_data_uri,
    }
    # Generate PDF using the utility function
    pdf = generate_pdf(form)
    
    # Create HTTP response with PDF
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="requirement_form_{request.user.username}_{form.process_name}.pdf"'
    
    return response

@user_passes_test(is_admin)
def admin_dashboard(request):
    # Basic counts
    total_users = User.objects.filter(role='user').count()
    total_admins = User.objects.filter(role='admin').count()
    submitted_forms = RequirementForm.objects.filter(status='submitted').count()
    draft_forms = RequirementForm.objects.filter(status='draft').count()
    
    # Calculate response rate
    response_rate = (submitted_forms / total_users) * 100 if total_users > 0 else 0
    
    # Count participating users (users who submitted at least one form)
    participating_users = User.objects.annotate(
        submission_count=Count('requirement_forms', filter=Q(requirement_forms__status='submitted'))
    ).filter(submission_count__gt=0).count()
    
    # Calculate user participation rate
    user_participation_rate = (participating_users / total_users) * 100 if total_users > 0 else 0
    
    # Count participating wings
    participating_wings = RequirementForm.objects.filter(
        status='submitted'
    ).values('user__wing_name').annotate(count=Count('user__wing_name')).count()
    
    # Get total wings count for participation rate
    total_wings = User.objects.values('wing_name').exclude(wing_name='').distinct().count()
    wing_participation_rate = (participating_wings / total_wings) * 100 if total_wings > 0 else 0
    
    # Forms by wing and department for charts
    forms_by_wing = RequirementForm.objects.filter(status='submitted').values('user__wing_name').annotate(count=Count('id'))
    forms_by_department = RequirementForm.objects.filter(status='submitted').values('user__department_name').annotate(count=Count('id'))
    
    # Get users by wing for the participating users chart
    users_by_wing = User.objects.values('wing_name').annotate(
        participating=Count('id', filter=Q(requirement_forms__status='submitted'), distinct=True)
    ).filter(participating__gt=0)

    # Prepare data for charts
    wing_labels = [item['user__wing_name'] or 'Unknown' for item in forms_by_wing]
    wing_data = [item['count'] for item in forms_by_wing]

    participating_wing_labels = [item['wing_name'] or 'Unknown' for item in users_by_wing]
    participating_wing_data = [item['participating'] for item in users_by_wing]
    
    context = {
        'total_users': total_users,
        'total_admins': total_admins,
        'submitted_forms': submitted_forms,
        'draft_forms': draft_forms,
        'response_rate': response_rate,
        'participating_users_count': participating_users,
        'user_participation_rate': user_participation_rate,
        'participating_wings_count': participating_wings,
        'wing_participation_rate': wing_participation_rate,
        'chart_data': {
            'wing_labels': wing_labels,
            'wing_data': wing_data,
            'participating_wing_labels': participating_wing_labels, 
            'participating_wing_data': participating_wing_data
        },
    }
    return render(request, 'requirements_app/admin_dashboard.html', context)


@user_passes_test(is_admin)
def user_management(request):
    # Get all unique wings for the filter dropdown
    wings = User.objects.values('wing_name').distinct().exclude(wing_name='').order_by('wing_name')
    
    # Base queryset
    users_list = User.objects.all()
    
    # Apply filters if provided
    filters_applied = False
    
    # Username filter
    username = request.GET.get('username', '')
    if username:
        users_list = users_list.filter(username__icontains=username)
        filters_applied = True
    
    # Name filter
    name = request.GET.get('name', '')
    if name:
        users_list = users_list.filter(
            Q(first_name__icontains=name) | 
            Q(last_name__icontains=name)
        )
        filters_applied = True
    
    # Wing filter
    wing = request.GET.get('wing', '')
    if wing:
        users_list = users_list.filter(wing_name=wing)
        filters_applied = True
    
    # Role filter
    role = request.GET.get('role', '')
    if role:
        users_list = users_list.filter(role=role)
        filters_applied = True
    
    # Order the results
    users_list = users_list.order_by('role', 'username')
    
    # Get total count before pagination
    total_users = User.objects.count()
    
    # Set up pagination
    paginator = Paginator(users_list, 10)  # Show 10 users per page
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    # Calculate visible page range
    page_range = []
    if paginator.num_pages <= 7:
        page_range = paginator.page_range
    else:
        current_page = page_obj.number
        if current_page <= 4:
            page_range = range(1, min(8, paginator.num_pages + 1))
        elif current_page >= paginator.num_pages - 3:
            page_range = range(max(1, paginator.num_pages - 6), paginator.num_pages + 1)
        else:
            page_range = range(current_page - 3, current_page + 4)
    
    context = {
        'users': users_list,  # Keep for backward compatibility
        'page_obj': page_obj,
        'page_range': page_range,
        'total_users': total_users,
        'filters_applied': filters_applied,
        'wings': wings,
    }
    return render(request, 'requirements_app/user_management.html', context)


@user_passes_test(is_admin)
def register_user(request):
    if request.method == 'POST':
        form = UserRegistrationForm(request.POST, request.FILES)
        if form.is_valid():
            # Create user but don't save to DB yet
            user = form.save(commit=False)
            user.set_password(form.cleaned_data['password'])
            
            # First save to generate ID
            user.save()
            
            # Process the uploaded files if they exist
            if 'photo' in request.FILES:
                # Save again to ensure the custom path with ID is used
                user.photo = request.FILES['photo']
                
            if 'signature' in request.FILES:
                user.signature = request.FILES['signature']
                
            # Save the final user object with all updates
            user.save()
            
            messages.success(request, f"User '{user.username}' created successfully.")
            return redirect('user_management')
    else:
        form = UserRegistrationForm()
    
    context = {
        'form': form,
    }
    return render(request, 'requirements_app/register_user.html', context)


@user_passes_test(is_admin)
def reset_password(request, user_id):
    user = get_object_or_404(User, pk=user_id)
    
    if request.method == 'POST':
        new_password = request.POST.get('new_password')
        user.set_password(new_password)
        user.save()
        messages.success(request, f"Password for '{user.username}' reset successfully.")
        return redirect('user_management')
    
    context = {
        'user': user,
    }
    return render(request, 'requirements_app/reset_password.html', context)

@user_passes_test(is_admin)
def change_role(request, user_id):
    user = get_object_or_404(User, pk=user_id)
    
    if request.method == 'POST':
        user.role = 'admin' if user.role == 'user' else 'user'
        user.save()
        messages.success(request, f"Role updated for '{user.username}' to {user.role}.")
    
    return redirect('user_management')

@user_passes_test(is_admin)
def view_responses(request):
    forms = RequirementForm.objects.filter(status='submitted').order_by('-submitted_at')
    
    # Filter by process name if provided
    process_filter = request.GET.get('process')
    if process_filter:
        forms = forms.filter(process_name__icontains=process_filter)
    
    #Check attachment existence for each form
    for form in forms:
        if form.attachment and form.attachment.name:
            form.has_attachment = True
        else:
            form.has_attachment = False
    # Get unique process names for filtering
    all_process_names = RequirementForm.objects.filter(status='submitted').values_list('process_name', flat=True).distinct()
    
    context = {
        'forms': forms,
        'all_process_names': all_process_names,
        'process_filter': process_filter,
    }
    return render(request, 'requirements_app/view_responses.html', context)


@user_passes_test(is_admin)
def delete_form(request, form_id):
    form = get_object_or_404(RequirementForm, pk=form_id)
    
    if request.method == 'POST':
        form.delete()
        messages.success(request, "Form deleted successfully.")
        return redirect('view_responses')
    
    context = {
        'form': form,
    }
    return render(request, 'requirements_app/delete_form_confirm.html', context)

@user_passes_test(is_admin)
def data_analysis(request):
    # Get some general statistics
    total_users = User.objects.filter(role='user').count()
    submitted_forms = RequirementForm.objects.filter(status='submitted').count()
    response_rate = (submitted_forms / total_users) * 100 if total_users > 0 else 0
    avg_completion_time = "5.2"  # This is a placeholder - calculate actual value if needed
    
    # Forms by wing and department
    forms_by_wing = RequirementForm.objects.filter(status='submitted').values('user__wing_name').annotate(count=Count('id'))
    forms_by_department = RequirementForm.objects.filter(status='submitted').values('user__department_name').annotate(count=Count('id'))
    
    # Process names
    process_names = RequirementForm.objects.filter(status='submitted').values('process_name').annotate(count=Count('id')).order_by('-count')[:8]
    
    # Prepare data for charts
    wing_labels = [item['user__wing_name'] or 'Unknown' for item in forms_by_wing]
    wing_data = [item['count'] for item in forms_by_wing]
    
    dept_labels = [item['user__department_name'] or 'Unknown' for item in forms_by_department]
    dept_data = [item['count'] for item in forms_by_department]
    
    process_names_data = [item['process_name'] for item in process_names]
    process_counts_data = [item['count'] for item in process_names]

    # Get all questions with responses
    questions = FormQuestion.objects.filter(
        id__in=QuestionResponse.objects.values_list('question', flat=True).distinct()
    )
    # Prepare question analysis data
    question_analysis = []
    for question in questions:
    
        # Get all responses for this question
        responses = QuestionResponse.objects.filter(question=question)
    
        # For choice-based questions (radio, select, checkbox)
        if question.field_type in ['radio', 'select', 'checkbox']:
            # Count occurrences of each option
            option_counts = {}
            for response in responses:
                # Split by comma for checkbox responses
                if question.field_type == 'checkbox' and ',' in response.response_text:
                    options = [opt.strip() for opt in response.response_text.split(',')]
                    for opt in options:
                        option_counts[opt] = option_counts.get(opt, 0) + 1
                else:
                    option_counts[response.response_text] = option_counts.get(response.response_text, 0) + 1
        
            # Calculate percentages
            total = sum(option_counts.values())
            options_data = []
            colors = ['primary', 'success', 'warning', 'danger', 'info', 'secondary']
        
            for i, (option, count) in enumerate(option_counts.items()):
                percentage = (count / total) * 100 if total > 0 else 0
                options_data.append({
                    'text': option,
                    'count': count,
                    'percentage': percentage,
                    'color': colors[i % len(colors)]
                })
        
            question_analysis.append({
                'text': question.question_text,
                'total_responses': total,
                'options': options_data
            })
    
        # For text-based questions, just count total responses
        else:
            question_analysis.append({
                'text': question.question_text,
                'total_responses': responses.count(),
                'options': []  # Empty options for text fields
            })
    
    # Prepare data for Time Efficiency Analysis chart
    time_vs_steps_data = []
    for form in RequirementForm.objects.filter(status='submitted'):
        if form.time_taken is not None and form.process_steps is not None:
            time_vs_steps_data.append({
                'x': form.process_steps,
                'y': form.time_taken,
                'process': form.process_name
            })

    # Prepare data for Process Steps Distribution
    steps_count = defaultdict(int)
    for form in RequirementForm.objects.filter(status='submitted'):
        if form.process_steps is not None:
            steps_count[form.process_steps] += 1

    # Convert to format for Chart.js
    steps_data = []
    for steps, count in sorted(steps_count.items()):
        steps_data.append({'x': steps, 'y': count})

    # Prepare Error Possibility data
    # error_ranges = {'0-20%': 0, '21-40%': 0, '41-60%': 0, '61-80%': 0, '81-100%': 0}
    # for form in RequirementForm.objects.filter(status='submitted'):
    #     if form.error_possibility is not None:
    #         if form.error_possibility <= 20:
    #             error_ranges['0-20%'] += 1
    #         elif form.error_possibility <= 40:
    #             error_ranges['21-40%'] += 1
    #         elif form.error_possibility <= 60:
    #             error_ranges['41-60%'] += 1
    #         elif form.error_possibility <= 80:
    #             error_ranges['61-80%'] += 1
    #         else:
    #             error_ranges['81-100%'] += 1

    # error_labels = list(error_ranges.keys())
    # error_data = list(error_ranges.values())
    
    context = {
    'question_analysis' : question_analysis,
    'total_users': total_users,
    'submitted_forms': submitted_forms,
    'response_rate': response_rate,
    'avg_completion_time': avg_completion_time,
    'chart_data': {
        'wing_labels': wing_labels,
        'wing_data': wing_data,
        'dept_labels': dept_labels,
        'dept_data': dept_data,
        'process_names': process_names_data,
        'process_counts_data': process_counts_data,
        # Add these with default values
        'time_vs_steps': [],
        'steps_data': []
        # 'error_labels': ['0-20%', '21-40%', '41-60%', '61-80%', '81-100%'],
        # 'error_data': [0, 0, 0, 0, 0]
    },
}
    return render(request, 'requirements_app/data_analysis.html', context)

@user_passes_test(is_admin)
def download_user_list(request):
    # Generate PDF of user list
    pdf = generate_user_list_pdf()
    
    # Create HTTP response with PDF
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="user_list.pdf"'
    
    return response

def login_view(request):
    if request.user.is_authenticated:
        if request.user.role == 'admin':
            return redirect('admin_dashboard')
        return redirect('dashboard')

    if request.method == 'POST':
        form = LoginForm(request.POST)
        if form.is_valid():
            username = form.cleaned_data['username']
            password = form.cleaned_data['password']
            user = authenticate(request, username=username, password=password)
            if user is not None:
                login(request, user)
                messages.success(request, f"Welcome back, {user.get_full_name()}!")
                if user.role == 'admin':
                    return redirect('admin_dashboard')
                return redirect('dashboard')
            else:
                messages.error(request, "Invalid username or password.")
    else:
        form = LoginForm()

    context = {
        'form': form,
    }
    return render(request, 'requirements_app/login.html', context)


@require_POST
def logout_view(request):
    logout(request)
    return redirect('login')

@user_passes_test(is_admin)
def edit_user(request, user_id):
    user = get_object_or_404(User, pk=user_id)
    
    if request.method == 'POST':
        form = UserUpdateForm(request.POST, request.FILES, instance=user)
        if form.is_valid():
            # Process the uploaded files if they exist
            if 'photo' in request.FILES:
                # Delete old photo if exists
                if user.photo:
                    if os.path.isfile(user.photo.path):
                        os.remove(user.photo.path)
                
                user.photo = request.FILES['photo']
                
            if 'signature' in request.FILES:
                # Delete old signature if exists
                if user.signature:
                    if os.path.isfile(user.signature.path):
                        os.remove(user.signature.path)
                
                user.signature = request.FILES['signature']
            
            form.save()
            messages.success(request, f"User '{user.username}' updated successfully.")
            return redirect('user_management')
    else:
        form = UserUpdateForm(instance=user)
    
    context = {
        'form': form,
        'user': user,
    }
    return render(request, 'requirements_app/edit_user.html', context)

@user_passes_test(is_admin)
def manage_form_structure(request):
    # Get all sections ordered by their order field
    sections = FormSection.objects.all().order_by('order')
    
    # For each section, manually add ordered questions
    for section in sections:
        section.ordered_questions = section.questions.all().order_by('order')
    
    context = {
        'sections': sections,
    }
    return render(request, 'requirements_app/manage_form_structure.html', context)

@user_passes_test(is_admin)
def edit_section(request, section_id=None):
    if section_id:
        section = get_object_or_404(FormSection, pk=section_id)
        title = f"Edit Section: {section.title}"
    else:
        section = None
        title = "Add New Section"
    
    if request.method == 'POST':
        form = FormSectionForm(request.POST, instance=section)
        if form.is_valid():
            form.save()
            messages.success(request, f"Section {'updated' if section else 'created'} successfully.")
            return redirect('manage_form_structure')
    else:
        form = FormSectionForm(instance=section)
    
    context = {
        'form': form,
        'title': title,
        'is_edit': section is not None,
    }
    return render(request, 'requirements_app/edit_section.html', context)

def delete_section(request, section_id):
    section = get_object_or_404(FormSection, pk=section_id)
    
    if request.method == 'POST':
        section.delete()
        messages.success(request, "Section deleted successfully.")
        return redirect('manage_form_structure')
    
    context = {
        'section': section,
    }
    return render(request, 'requirements_app/delete_section_confirm.html', context)

@user_passes_test(is_admin)
def edit_question(request, section_id, question_id=None):
    section = get_object_or_404(FormSection, pk=section_id)
    
    if question_id:
        question = get_object_or_404(FormQuestion, pk=question_id, section=section)
        title = f"Edit Question: {question.question_text}"
    else:
        question = None
        title = f"Add New Question to {section.title}"
    
    if request.method == 'POST':
        form = FormQuestionForm(request.POST, instance=question)
        if form.is_valid():
            new_question = form.save(commit=False)
            new_question.section = section
            new_question.save()
            
            messages.success(request, f"Question {'updated' if question_id else 'created'} successfully.")
            return redirect('manage_form_structure')
        else:
            # Display form errors
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
    else:
        # For a new question, set default values
        initial_data = {}
        if not question:
            # Get the highest order value in this section and add 1
            highest_order = FormQuestion.objects.filter(section=section).order_by('-order').values_list('order', flat=True).first()
            initial_data['order'] = (highest_order or 0) + 1
            initial_data['is_active'] = True
            
        form = FormQuestionForm(instance=question, initial=initial_data)
    
    context = {
        'form': form,
        'section': section,
        'title': title,
        'is_edit': question is not None,
    }
    return render(request, 'requirements_app/edit_question.html', context)

@user_passes_test(is_admin)
def delete_question(request, question_id):
    question = get_object_or_404(FormQuestion, pk=question_id)
    section = question.section
    
    if request.method == 'POST':
        question.delete()
        messages.success(request, "Question deleted successfully.")
        return redirect('manage_form_structure')
    
    context = {
        'question': question,
        'section': section,
    }
    return render(request, 'requirements_app/delete_question_confirm.html', context)

@user_passes_test(is_admin)
def download_user_pdf(request, form_id):
    form = get_object_or_404(RequirementForm, pk=form_id)
    
    # Get all sections with questions and answers (same as view_user_form)
    sections = FormSection.objects.filter(is_active=True).order_by('order')
    sections_with_answers = []
    for section in sections:
        section_answers = []
        for question in section.questions.filter(is_active=True):
            try:
                answer = QuestionResponse.objects.get(form=form, question=question)
                section_answers.append({
                    'question': question,
                    'answer_text': answer.response_text
                })
            except QuestionResponse.DoesNotExist:
                continue
        if section_answers:
            sections_with_answers.append({
                'section': section,
                'answers': section_answers
            })

    # Generate the signature data URI
    signature_data_uri = None
    if form.user.signature:  # Check if the user has a signature
        signature_path = os.path.join(settings.MEDIA_ROOT, form.user.signature.name)
        if os.path.exists(signature_path):
            with open(signature_path, 'rb') as image_file:
                signature_data = base64.b64encode(image_file.read()).decode('utf-8')
                signature_data_uri = f"data:image/png;base64,{signature_data}"  # Adjust MIME type if needed
        else:
            print(f"Signature file not found at: {signature_path}")

    # Prepare context for the template
    context = {
        'form': form,
        'sections_with_answers': sections_with_answers,
        'user': form.user,  # Use the form's user, not the request user
        'current_date': timezone.now(),
        'signature_data_uri': signature_data_uri,
    }
    
    # Generate PDF using the utility function
    pdf = generate_pdf(form)
    
    # Create HTTP response with PDF
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="requirement_form_{form.user.username}.pdf"'
    
    return response

@user_passes_test(is_admin)
def delete_attachment(request, form_id):
    form = get_object_or_404(RequirementForm, pk=form_id)
    
    if request.method == 'POST':
        # Check if attachment exists
        if form.attachment:
            # Delete the file from storage
            if os.path.isfile(form.attachment.path):
                os.remove(form.attachment.path)
            
            # Clear the field
            form.attachment = None
            form.save()
            
            messages.success(request, "Attachment deleted successfully.")
        else:
            messages.warning(request, "No attachment found.")
            
        return redirect('view_responses')
    
    context = {
        'form': form,
    }
    return render(request, 'requirements_app/delete_attachment_confirm.html', context)

@login_required
def delete_draft(request, form_id):
    draft = get_object_or_404(RequirementForm, id=form_id, user=request.user, status='draft')
    
    if request.method == 'POST':
        draft.delete()
        messages.success(request, "Draft form deleted successfully.")
        return redirect('dashboard')
    
    context = {
        'draft': draft,
    }
    return render(request, 'requirements_app/delete_draft_confirm.html', context)

@login_required
def review_form(request):
    # Check if there's form data in session
    if 'form_data' not in request.session or 'question_responses' not in request.session:
        messages.error(request, "No form data found for review. Please fill out the form again.")
        return redirect('form')
    
    # Get form data from session
    form_data = request.session['form_data']
    question_responses = request.session['question_responses']
    
    # Add this code to ensure process_steps_detail is properly formatted
    process_steps_detail = form_data.get('process_steps_detail', [])
    
    # For debugging
    print("Type of process_steps_detail:", type(process_steps_detail))
    print("Content of process_steps_detail:", process_steps_detail)
    
    # Ensure it's properly parsed if it's a string
    if isinstance(process_steps_detail, str):
        try:
            process_steps_detail = json.loads(process_steps_detail)
            form_data['process_steps_detail'] = process_steps_detail
        except json.JSONDecodeError:
            form_data['process_steps_detail'] = []
    
    # Create a mock form object to mimic the structure expected by the template
    form = type('obj', (object,), {
        'process_name': form_data.get('process_name', ''),
        'process_description': form_data.get('process_description', ''),
        'process_steps_detail': form_data.get('process_steps_detail', []),
        'has_flowchart': form_data.get('has_flowchart', False),
        'has_attachment': form_data.get('has_attachment', False),
        # Add other fields as needed
    })
    
    # Get active sections
    active_sections = FormSection.objects.filter(is_active=True).prefetch_related('questions').order_by('order')
    
    # Create structure for template rendering
    sections_with_answers = []
    for section in active_sections:
        section_answers = []
        
        for question in section.questions.filter(is_active=True).order_by('order'):
            # Check if we have a response for this question
            if str(question.id) in question_responses or question.id in question_responses:
                # Get answer for this question
                question_id = str(question.id) if str(question.id) in question_responses else question.id
                answer_text = question_responses[question_id]
                
                section_answers.append({
                    'question': question,
                    'answer_text': answer_text
                })
        
        # Only add sections with answers
        if section_answers:
            sections_with_answers.append({
                'section': section,
                'answers': section_answers
            })
    
    context = {
        'form_data': form_data,
        'form': form,  # Add this to provide the template with the expected structure
        'sections_with_answers': sections_with_answers
    }
    
    return render(request, 'requirements_app/review_form.html', context)

@login_required
def edit_form(request):
    # Check if user came from review page
    if 'form_data' not in request.session:
        messages.error(request, "No form data found. Please fill out the form again.")
        return redirect('form')
    
    # Get form_id if we're editing an existing draft
    form_data = request.session['form_data']
    form_id = form_data.get('form_id')
    
    # Store an edit flag in the session
    request.session['edit_after_review'] = True
    
    # Redirect back to form page
    if form_id:
        return redirect('form_view', form_id=form_id)
    else:
        return redirect('form')

@login_required
def confirm_submission(request):
    # Check if user came from review page
    if 'form_data' not in request.session or 'question_responses' not in request.session:
        messages.error(request, "No form data found. Please fill out the form again.")
        return redirect('form')
    
    if request.method != 'POST':
        return redirect('review_form')
    
    # Get data from session
    form_data = request.session['form_data']
    question_responses = request.session['question_responses']
    temp_files = request.session.get('temp_files', {})
    
    # Get or create form instance
    form_id = form_data.get('form_id')
    if form_id:
        form_instance = get_object_or_404(RequirementForm, id=form_id, user=request.user, status='draft')
    else:
        form_instance = RequirementForm(user=request.user)
    
    # Update form instance with data from form_data
    for field, value in form_data.items():
        if field in ['form_id', 'has_flowchart', 'has_attachment']:
            continue
        setattr(form_instance, field, value)
    
    # Set status to submitted and save timestamp
    form_instance.status = 'submitted'
    form_instance.submitted_at = timezone.now()
    
    # Handle file uploads from temp files
    if 'flowchart' in temp_files:
        with open(temp_files['flowchart'], 'rb') as f:
            content = f.read()
            filename = f"{request.user.username}_DataFlow.pdf"
            form_instance.flowchart.save(filename, ContentFile(content), save=False)
    
    if 'attachment' in temp_files:
        with open(temp_files['attachment'], 'rb') as f:
            content = f.read()
            filename = f"{request.user.username}_{timezone.now().strftime('%Y%m%d%H%M%S')}.pdf"
            form_instance.attachment.save(filename, ContentFile(content), save=False)
    
    # Save the form
    form_instance.save()
    
    # Save question responses
    for question_id, answer_text in question_responses.items():
        try:
            question = FormQuestion.objects.get(id=int(question_id))
            QuestionResponse.objects.update_or_create(
                form=form_instance,
                question=question,
                defaults={'response_text': answer_text}
            )
        except (FormQuestion.DoesNotExist, ValueError):
            continue
    
    # Clean up temp files
    for temp_path in temp_files.values():
        try:
            os.unlink(temp_path)
        except (OSError, IOError):
            pass
    
    # Clean up session
    for key in ['form_data', 'question_responses', 'temp_files', 'review_timestamp']:
        if key in request.session:
            del request.session[key]
    
    # Show success message
    messages.success(request, "Form submitted successfully!")
    
    # Redirect to view submission page
    return redirect('view_submissions')


def generate_report_pdf(template_path, context, filename="report.pdf"):
    """Generate PDF report from a template and context data"""
    # Render HTML string from template
    html_string = render_to_string(template_path, context)
    
    # Create a temporary file
    with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as output:
        # Generate PDF
        HTML(string=html_string).write_pdf(
            output,
            stylesheets=[
                CSS(string='@page { size: A4; margin: 1cm }')
            ]
        )
        output_path = output.name
    
    # Read the generated PDF
    with open(output_path, 'rb') as f:
        pdf_content = f.read()
    
    # Clean up the temporary file
    os.unlink(output_path)
    
    # Create HTTP response with PDF
    response = HttpResponse(pdf_content, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response

@user_passes_test(is_admin)
def edit_submitted_form(request, form_id):
    # Get the form instance
    form_instance = get_object_or_404(RequirementForm, pk=form_id, status='submitted')
    
    # Get all active form sections with their questions
    active_sections = FormSection.objects.filter(is_active=True).prefetch_related('questions').order_by('order')

    if request.method == 'POST':
        # Process the basic form data
        basic_form = RequirementFormForm(request.POST, request.FILES, instance=form_instance)
        # Process the dynamic questions
        dynamic_form = DynamicForm(request.POST, sections=active_sections, instance=form_instance)

        if basic_form.is_valid() and dynamic_form.is_valid():
            # Save the basic form
            requirement_form = basic_form.save(commit=False)
            
            # Process steps handling
            process_steps_detail = request.POST.get('process_steps_detail', '[]')
            try:
                steps_data = json.loads(process_steps_detail)
                valid_steps = []
                for i, step in enumerate(steps_data):
                    if isinstance(step, dict) and 'description' in step and step['description'].strip():
                        step['number'] = step.get('number', i + 1)
                        valid_steps.append(step)
                    elif isinstance(step, str) and step.strip():
                        valid_steps.append({
                            'number': i + 1,
                            'description': step.strip()
                        })
                
                requirement_form.process_steps_detail = valid_steps
                requirement_form.process_steps = len(valid_steps)
            except json.JSONDecodeError:
                requirement_form.process_steps_detail = []
                requirement_form.process_steps = 0
            
            # Handle file uploads
            if 'attachment' in request.FILES:
                file = request.FILES['attachment']
                if file.name.endswith('.pdf'):
                    filename = f"{form_instance.user.username}_{timezone.now().strftime('%Y%m%d%H%M%S')}.pdf"
                    content = file.read()
                    requirement_form.attachment.save(filename, ContentFile(content), save=False)
                else:
                    messages.error(request, "Attachment must be a PDF file.")
                    return redirect('edit_submitted_form', form_id=form_id)
                    
            # Handle flowchart upload (if provided)
            if 'flowchart' in request.FILES:
                file = request.FILES['flowchart']
                if file.name.endswith('.pdf'):
                    # Model will handle filename with custom upload path
                    pass
                else:
                    messages.error(request, "Flowchart must be a PDF file.")
                    return redirect('edit_submitted_form', form_id=form_id)
            
            # Save the form
            requirement_form.updated_at = timezone.now()
            requirement_form.save()

            # Save responses for dynamic questions
            for field_name, value in dynamic_form.cleaned_data.items():
                if field_name.startswith('question_'):
                    question_id = int(field_name.split('_')[1])
                    question = FormQuestion.objects.get(id=question_id)

                    # Convert list values to string for checkboxes
                    if isinstance(value, list):
                        value = ', '.join(value)

                    # Create or update response
                    QuestionResponse.objects.update_or_create(
                        form=requirement_form,
                        question=question,
                        defaults={'response_text': value if value is not None else ''}
                    )

            messages.success(request, "Form updated successfully.")
            return redirect('view_responses')
        else:
            # If forms are invalid, display errors
            for field, errors in basic_form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")

            for field, errors in dynamic_form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")

    else:
        # Initial form load
        basic_form = RequirementFormForm(instance=form_instance)
        dynamic_form = DynamicForm(sections=active_sections, instance=form_instance)
        
    # Group questions by section (same as in form_view)
    sections_with_questions = []
    for section in active_sections:
        questions = []
        for question in section.questions.filter(is_active=True).order_by('order'):
            field_name = f"question_{question.id}"
            if field_name in dynamic_form.fields:
                questions.append({
                    'question': question,
                    'field': dynamic_form[field_name]
                })

        if questions:  # Only add sections with active questions
            sections_with_questions.append({
                'section': section,
                'questions': questions
            })

    context = {
        'basic_form': basic_form,
        'dynamic_form': dynamic_form,
        'sections_with_questions': sections_with_questions,
        'form_instance': form_instance,
        'form_id': form_id,
    }

    return render(request, 'requirements_app/edit_submitted_form.html', context)

@login_required
def download_review_pdf(request):
    """Generate a PDF of the form being reviewed"""
    # Check if there's form data in session
    if 'form_data' not in request.session or 'question_responses' not in request.session:
        messages.error(request, "No form data found for review. Please fill out the form again.")
        return redirect('form')
    
    # Get form data from session
    form_data = request.session['form_data']
    question_responses = request.session['question_responses']
    
    # Create a mock form object to mimic the structure expected by the generate_pdf function
    mock_form = type('obj', (object,), {
        'user': request.user,
        'process_name': form_data.get('process_name', ''),
        'process_description': form_data.get('process_description', ''),
        'process_steps_detail': form_data.get('process_steps_detail', []),
        'time_taken': form_data.get('time_taken', ''),
        'people_involved': form_data.get('people_involved', ''),
        'process_steps': form_data.get('process_steps', ''),
        'expected_features': form_data.get('expected_features', ''),
        'internal_connectivity': form_data.get('internal_connectivity', ''),
        'internal_connectivity_details': form_data.get('internal_connectivity_details', ''),
        'external_connectivity': form_data.get('external_connectivity', ''),
        'external_connectivity_details': form_data.get('external_connectivity_details', ''),
        'expected_reports': form_data.get('expected_reports', ''),
        'expected_analysis': form_data.get('expected_analysis', ''),
        'status': 'draft',
        'created_at': timezone.now(),
        'updated_at': timezone.now(),
        'submitted_at': None,
    })
    
    # Get all active sections with questions
    active_sections = FormSection.objects.filter(is_active=True).order_by('order')
    
    # Create a custom generate_preview_pdf function or modify the existing one
    # Since the form is not yet saved, we need to handle this differently
    
    # Get signature data URI
    signature_data_uri = None
    if request.user.signature:
        try:
            with open(request.user.signature.path, 'rb') as f:
                signature_content = f.read()
                _, ext = os.path.splitext(request.user.signature.path)
                mime_type = {
                    '.png': 'image/png',
                    '.jpg': 'image/jpeg',
                    '.jpeg': 'image/jpeg',
                    '.gif': 'image/gif'
                }.get(ext.lower(), 'image/jpeg')
                signature_data_uri = f"data:{mime_type};base64,{base64.b64encode(signature_content).decode('utf-8')}"
        except Exception as e:
            print(f"Error processing signature: {str(e)}")
    
    # Get logo as data URI
    logo_data_uri = get_static_file_as_data_uri('images/parliament_logo.png')
    
    # Create a map for storing question responses
    responses_map = {}
    for question_id, response_text in question_responses.items():
        try:
            question = FormQuestion.objects.get(id=int(question_id))
            responses_map[question.id] = response_text
        except (FormQuestion.DoesNotExist, ValueError):
            continue
    
    # Create sections with answers structure
    sections_with_answers = []
    for section in active_sections:
        section_answers = []
        for question in section.questions.filter(is_active=True).order_by('order'):
            if question.id in responses_map:
                section_answers.append({
                    'question': question,
                    'answer_text': responses_map[question.id]
                })
        
        if section_answers:
            sections_with_answers.append({
                'section': section,
                'answers': section_answers
            })
    
    # Render HTML content
    html_string = render_to_string('requirements_app/pdf/form_template.html', {
        'form': mock_form,
        'user': request.user,
        'logo_data_uri': logo_data_uri,
        'signature_data_uri': signature_data_uri,
        'current_date': timezone.now(),
        'sections_with_answers': sections_with_answers
    })
    
    # Create a temporary file
    with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as output:
        # Generate PDF
        HTML(string=html_string).write_pdf(
            output,
            stylesheets=[
                CSS(string='@page { size: A4; margin: 1.5cm }')
            ]
        )
        output_path = output.name
    
    # Read the generated PDF
    with open(output_path, 'rb') as f:
        pdf_content = f.read()
    
    # Clean up the temporary file
    os.unlink(output_path)
    
    # Create HTTP response with PDF
    response = HttpResponse(pdf_content, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="requirement_form_preview_{request.user.username}.pdf"'
    
    return response


@user_passes_test(is_admin)
def reports(request):
    context = {}
    
    # Get report type from request
    report_type = request.GET.get('report_type', '')
    context['report_type'] = report_type
    
    # 1. System User Reports
    if report_type == 'users':
        user_role = request.GET.get('user_role', 'all')
        context['user_role'] = user_role
        
        # Query users based on role
        if user_role == 'admin':
            user_report = User.objects.filter(role='admin').order_by('username')
            context['user_report_title'] = 'Admin Users List'
        elif user_role == 'user':
            user_report = User.objects.filter(role='user').order_by('username')
            context['user_report_title'] = 'Regular Users List'
        else:
            user_report = User.objects.all().order_by('role', 'username')
            context['user_report_title'] = 'All System Users List'
        
        context['user_report'] = user_report
    
    # 2. Submitted Forms Reports
    elif report_type == 'forms':
        form_filter_type = request.GET.get('form_filter_type', 'wing')
        context['form_filter_type'] = form_filter_type
        
        # Get all available wings for dropdown
        available_wings = User.objects.exclude(wing_name='').values('wing_name').distinct().order_by('wing_name')
        context['available_wings'] = available_wings
        
        # A. Wing-wise report
        if form_filter_type == 'wing':
            selected_wings = request.GET.getlist('selected_wings', [])
            context['selected_wings'] = selected_wings
            
            # Query forms based on selected wings
            if not selected_wings or 'all' in selected_wings:
                forms_report = RequirementForm.objects.filter(status='submitted').order_by('user__wing_name', '-submitted_at')
                context['forms_report_title'] = 'All Wings - Submitted Forms Report'
            else:
                forms_report = RequirementForm.objects.filter(status='submitted', user__wing_name__in=selected_wings).order_by('user__wing_name', '-submitted_at')
                context['forms_report_title'] = f"Selected Wings - Submitted Forms Report ({', '.join(selected_wings)})"
            
            context['forms_report'] = forms_report
        
        # B. User-wise report
        elif form_filter_type == 'user':
            selected_users = request.GET.get('selected_users', 'all')
            context['selected_users'] = selected_users
            
            if selected_users == 'all':
                forms_report = RequirementForm.objects.filter(status='submitted').order_by('user__username', '-submitted_at')
                context['forms_report_title'] = 'All Users - Submitted Forms Report'
            elif selected_users == 'custom':
                custom_user_ids = request.GET.get('custom_user_ids', '')
                context['custom_user_ids'] = custom_user_ids
                
                if custom_user_ids:
                    user_id_list = [user_id.strip() for user_id in custom_user_ids.split(',')]
                    forms_report = RequirementForm.objects.filter(status='submitted', user__username__in=user_id_list).order_by('user__username', '-submitted_at')
                    context['forms_report_title'] = f"Selected Users - Submitted Forms Report ({custom_user_ids})"
                else:
                    forms_report = []
                    context['forms_report_title'] = 'No Users Selected'
            
            context['forms_report'] = forms_report
        
        # C. Process-wise report
        elif form_filter_type == 'process':
            process_name = request.GET.get('process_name', '')
            context['process_name'] = process_name
            
            if process_name:
                process_list = [p.strip() for p in process_name.split(',')]
                if len(process_list) == 1:
                    forms_report = RequirementForm.objects.filter(status='submitted', process_name__icontains=process_list[0]).order_by('process_name', '-submitted_at')
                    context['forms_report_title'] = f"Process: {process_name} - Submitted Forms Report"
                else:
                    q_objects = Q()
                    for process in process_list:
                        q_objects |= Q(process_name__icontains=process)
                    forms_report = RequirementForm.objects.filter(status='submitted').filter(q_objects).order_by('process_name', '-submitted_at')
                    context['forms_report_title'] = f"Multiple Processes - Submitted Forms Report ({process_name})"
            else:
                forms_report = []
                context['forms_report_title'] = 'No Process Selected'
            
            context['forms_report'] = forms_report
    
    # 3. Counting Based Reports
    elif report_type == 'counts':
        count_report_type = request.GET.get('count_report_type', 'wing_count')
        context['count_report_type'] = count_report_type
        
        # A. Wing-wise count report
        if count_report_type == 'wing_count':
            wing_count_report = RequirementForm.objects.filter(status='submitted').values('user__wing_name').annotate(count=Count('id')).order_by('user__wing_name')
            total_requirements = RequirementForm.objects.filter(status='submitted').count()
            
            context['wing_count_report'] = wing_count_report
            context['total_requirements'] = total_requirements
        
        # B. User participation count report
        elif count_report_type == 'user_participation':
            participation_status = request.GET.get('participation_status', 'all')
            context['participation_status'] = participation_status
            
            # Annotate users with submission count
            users = User.objects.filter(role='user').annotate(
                submission_count=Coalesce(Count('requirement_forms', filter=Q(requirement_forms__status='submitted')), V(0))
            )
            
            # Filter based on participation status
            if participation_status == 'submitted':
                users = users.filter(submission_count__gt=0)
            elif participation_status == 'not_submitted':
                users = users.filter(submission_count=0)
            
            # Sort users by wing name and username
            users = users.order_by('wing_name', 'username')
            
            context['user_participation_report'] = users
    
    return render(request, 'requirements_app/reports.html', context)

# this function for Excel download
@user_passes_test(is_admin)
def download_excel_report(request):
    # Get user IDs from the request
    user_ids = request.GET.get('user_ids', '')
    if not user_ids:
        messages.error(request, "No users selected for report generation")
        return redirect('reports')
    
    # Parse user IDs
    user_id_list = [uid.strip() for uid in user_ids.split(',')]
    
    # Create a new workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "E-Parliament Requirements"
    
    # Define styles
    header_font = Font(name='Calibri', bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='28A745', end_color='28A745', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # Set column widths
    column_widths = {
        'A': 5,   # SL. No.
        'B': 12,  # ID
        'C': 20,  # Name
        'D': 15,  # Wing
        'E': 20,  # Process Name
        'F': 30,  # Process Description
        'G': 20,  # Process Steps
        'H': 10,  # Time Taken
        'I': 10,  # People Involved
        'J': 10,  # Process Steps Count
        'K': 30,  # Expected Features
        'L': 10,  # Internal Connectivity
        'M': 30,  # Internal Connectivity Details
        'N': 10,  # External Connectivity
        'O': 30,  # Expected Reports
        'P': 30,  # Expected Analysis
        'Q': 30,  # Additional Comments
        'R': 15,  # Submitted At
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Define headers
    headers = [
        "SL.", "ID", "Name", "Wing", 
        "Process Name", "Process Description", "Process Steps", 
        "Time Taken", "People Involved", "Steps Count",
        "Expected Features", "Internal Connectivity", "Internal Connectivity Details", 
        "External Connectivity", "Expected Reports", "Expected Analysis",
        "Additional Comments", "Submitted At"
    ]
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Get the form data for the selected users
    if 'all' in user_id_list:
        forms = RequirementForm.objects.filter(status='submitted').order_by('user__username')
    else:
        forms = RequirementForm.objects.filter(
            status='submitted', 
            user__username__in=user_id_list
        ).order_by('user__username')
    
    # Write data rows
    row_num = 2
    for i, form in enumerate(forms, 1):
        # Get additional question responses
        additional_comments = ""
        try:
            # Find a response that might contain additional comments
            # This is just a placeholder - adjust based on your actual data structure
            additional_response = QuestionResponse.objects.filter(
                form=form,
                question__question_text__icontains='comment'
            ).first()
            
            if additional_response:
                additional_comments = additional_response.response_text
        except:
            pass
        
        # Format process steps data
        process_steps_text = ""
        if form.process_steps_detail:
            try:
                if isinstance(form.process_steps_detail, list):
                    steps = []
                    for step in form.process_steps_detail:
                        if isinstance(step, dict) and 'description' in step:
                            steps.append(f"{step.get('number', '')}: {step['description']}")
                        elif isinstance(step, str):
                            steps.append(step)
                    process_steps_text = "\n".join(steps)
            except:
                process_steps_text = str(form.process_steps_detail)
        
        # Row data
        row = [
            i,  # SL.
            form.user.username,  # ID
            form.user.get_full_name(),  # Name
            form.user.wing_name,  # Wing
            form.process_name,  # Process Name
            form.process_description,  # Process Description
            process_steps_text,  # Process Steps
            form.time_taken,  # Time Taken
            form.people_involved,  # People Involved
            form.process_steps,  # Steps Count
            form.expected_features,  # Expected Features
            form.get_internal_connectivity_display(),  # Internal Connectivity
            form.internal_connectivity_details if form.internal_connectivity == 'yes' else "",  # Internal Connectivity Details
            form.get_external_connectivity_display(),  # External Connectivity
            form.expected_reports,  # Expected Reports
            form.expected_analysis,  # Expected Analysis
            additional_comments,  # Additional Comments
            form.submitted_at.strftime('%Y-%m-%d %H:%M') if form.submitted_at else ""  # Submitted At
        ]
        
        # Write row data
        for col_num, cell_value in enumerate(row, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.border = border
            
            # Set wrap text for text columns
            if col_num in [5, 6, 7, 11, 13, 15, 16, 17]:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        row_num += 1
    
    # Freeze the header row
    ws.freeze_panes = 'A2'
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=E-Parliament_Requirements_Report.xlsx'
    
    # Save the workbook to the response
    wb.save(response)
    
    return response